import random
import re
import string
from collections import defaultdict
import openpyxl
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from django.forms import inlineformset_factory
from django.http import HttpResponseForbidden, HttpResponseNotFound, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import UpdateView
from tabeasy.settings import DEBUG
from django.core.files.storage import FileSystemStorage

from accounts.models import User
from submission.forms import CharacterPronounsForm
from submission.models.captains_meeting import CaptainsMeeting
from submission.models.paradigm import Paradigm, ParadigmPreferenceItem, ParadigmPreference, \
    experience_description_choices
from submission.models.section import Section, SubSection
from submission.models.spirit import Spirit
from tabeasy.settings import DEBUG
from tabeasy.utils.mixins import JudgeOnlyMixin, PassRequestToFormViewMixin, TabOnlyMixin
from tourney.forms import RoundForm, UpdateConflictForm, UpdateJudgeFriendForm, PairingFormSet, PairingSubmitForm, \
    JudgeForm, CheckinJudgeForm, CompetitorPronounsForm, TournamentForm, CreateTournamentForm, CompetitorForm, TeamForm, \
    ByebusterGenerateForm, GlobalScoringJudgeForm
from submission.models.ballot import Ballot
from submission.models.character import Character, CharacterPronouns
from tourney.models import Tournament
from tourney.models.judge import Judge
from tourney.models.round import Round, Pairing
from tourney.models.team import Team
from tourney.models.competitor import Competitor


def get_team_competitor_formset(tournament):
    competitor_slots = max(1, tournament.team_size)
    return inlineformset_factory(
        Team,
        Competitor,
        fields=('name', 'pronouns'),
        max_num=competitor_slots,
        validate_max=True,
        extra=competitor_slots,
    )


def lock_form_fields(form):
    for field in form.fields.values():
        field.disabled = True


def set_pairing_banner(request, tournament, errors, **extra):
    request.session['extra'] = {
        'errors': errors,
        'tournament_id': tournament.pk,
        **extra,
    }


def normalize_import_username(username):
    return re.sub(r'[^a-zA-Z0-9_-]+', '_', str(username or '').strip()).strip('_').lower()


def normalize_import_value(value):
    return str(value or '').strip()


def normalize_excel_header(value):
    value = re.sub(r'[^a-zA-Z0-9]+', ' ', str(value or '').strip().lower())
    return re.sub(r'\s+', ' ', value).strip()


def header_matches(value, expected):
    value = normalize_excel_header(value)
    expected = normalize_excel_header(expected)
    return value == expected or value.startswith(f'{expected} ')


def find_judge_header_row(worksheet):
    for row_num in range(1, min(worksheet.max_row, 10) + 1):
        values = [normalize_excel_header(worksheet.cell(row_num, col).value) for col in range(1, worksheet.max_column + 1)]
        if 'first name' in values and 'last name' in values:
            return row_num
    return 1


def find_header_column(worksheet, header_row, expected):
    for col in range(1, worksheet.max_column + 1):
        if header_matches(worksheet.cell(header_row, col).value, expected):
            return col
    return None


def append_header_column(worksheet, header_row, label):
    col = worksheet.max_column + 1
    worksheet.cell(row=header_row, column=col).value = label
    return col


def get_judge_sheet_columns(tournament, worksheet, create_missing=False):
    header_row = find_judge_header_row(worksheet)
    first_name_col = find_header_column(worksheet, header_row, 'First Name') or 1
    last_name_col = find_header_column(worksheet, header_row, 'Last Name') or 2
    preside_col = find_header_column(worksheet, header_row, 'Presiding') or 3
    username_col = find_header_column(worksheet, header_row, 'Username')
    password_col = find_header_column(worksheet, header_row, 'Password')
    if create_missing and not username_col:
        username_col = append_header_column(worksheet, header_row, 'Username')
    if create_missing and not password_col:
        password_col = append_header_column(worksheet, header_row, 'Password')

    round_columns = {}
    for round_num in range(1, min(tournament.total_rounds, 9) + 1):
        round_columns[round_num] = find_header_column(worksheet, header_row, tournament.get_round_label(round_num))

    return {
        'header_row': header_row,
        'first_name': first_name_col,
        'last_name': last_name_col,
        'preside': preside_col,
        'username': username_col,
        'password': password_col,
        'rounds': round_columns,
    }


def get_import_username_for_tournament(tournament, username):
    base = normalize_import_username(username)
    tournament_prefix = normalize_import_username(tournament.short_name or tournament.name or 'tournament')
    if not base:
        base = 'account'
    if User.objects.filter(username=base, tournament=tournament).exists():
        return base
    if not User.objects.filter(username=base).exists():
        return base

    localized_base = f'{tournament_prefix}_{base}' if tournament_prefix else base
    candidate = localized_base
    index = 2
    while User.objects.filter(username=candidate).exists():
        candidate = f'{localized_base}_{index}'
        index += 1
    return candidate


def get_judge_import_username(tournament, first_name, last_name):
    first_name = normalize_import_username(first_name)
    last_name = normalize_import_username(last_name)
    tournament_prefix = normalize_import_username(tournament.short_name or tournament.name or 'tournament')
    return f'{tournament_prefix}_{first_name}_{last_name}'.strip('_')


def get_legacy_judge_import_username(first_name, last_name):
    return f'{normalize_import_username(first_name)}_{normalize_import_username(last_name)}'.strip('_')


def get_judge_import_name_key(first_name, last_name):
    return (
        re.sub(r'\s+', ' ', normalize_import_value(first_name)).lower(),
        re.sub(r'\s+', ' ', normalize_import_value(last_name)).lower(),
    )


def find_existing_import_user(tournament, username, first_name, last_name):
    usernames = [
        normalize_import_username(username),
        get_judge_import_username(tournament, first_name, last_name),
        get_legacy_judge_import_username(first_name, last_name),
    ]
    seen_usernames = set()
    for candidate in usernames:
        if not candidate or candidate in seen_usernames:
            continue
        seen_usernames.add(candidate)
        user = User.objects.filter(username=candidate).first()
        if user:
            return user

    first_name_key, last_name_key = get_judge_import_name_key(first_name, last_name)
    if not first_name_key:
        return None
    for user in User.objects.filter(tournament=tournament, is_judge=True):
        if get_judge_import_name_key(user.first_name, user.last_name) == (first_name_key, last_name_key):
            return user
    return None


def get_user_judge(user):
    try:
        return user.judge
    except Judge.DoesNotExist:
        return None


try:
    from tabeasy_secrets.secret import str_int
except ImportError:
    str_int = int


def sort_teams(teams):
    return list(reversed(sorted(teams,
                                key=lambda x: (x.total_ballots, x.total_cs, x.total_pd))))


def sorted_break_teams(tournament, teams=None):
    queryset = teams if teams is not None else Team.objects.filter(user__tournament=tournament)
    teams = list(queryset)
    if teams and all(team.prelim_seed for team in teams):
        return sorted(teams, key=lambda team: team.prelim_seed)
    return sort_teams(teams)


def counted_ballots_for_round(round_obj):
    return round_obj.sync_counted_ballots()


def get_prelim_stats(team, prelim_rounds):
    ballots = 0
    pd = 0
    rounds = list(team.p_rounds.filter(pairing__round_num__lte=prelim_rounds)) + list(
        team.d_rounds.filter(pairing__round_num__lte=prelim_rounds)
    )
    for round_obj in rounds:
        round_ballots = [
            ballot for ballot in counted_ballots_for_round(round_obj)
            if ballot.byebuster_excluded_team_id != team.pk
        ]
        if round_obj.p_team == team:
            ballots += sum(ballot.p_ballot for ballot in round_ballots)
            pd += sum(ballot.p_pd for ballot in round_ballots)
        else:
            ballots += sum(ballot.d_ballot for ballot in round_ballots)
            pd += sum(ballot.d_pd for ballot in round_ballots)
    return ballots, pd


def rank_teams_for_break(tournament, teams=None):
    if teams is None:
        teams = sorted_break_teams(tournament)
        if teams and all(team.prelim_seed for team in teams):
            return [
                (
                    team,
                    team.locked_prelim_ballots if team.locked_prelim_ballots is not None else team.total_ballots,
                    team.locked_prelim_pd if team.locked_prelim_pd is not None else team.total_pd,
                )
                for team in teams
            ]
    else:
        teams = sorted_break_teams(tournament, teams)
        if teams and all(team.prelim_seed for team in teams):
            return [
                (
                    team,
                    team.locked_prelim_ballots if team.locked_prelim_ballots is not None else team.total_ballots,
                    team.locked_prelim_pd if team.locked_prelim_pd is not None else team.total_pd,
                )
                for team in teams
            ]
    ranked = []
    for team in teams:
        ballots, pd = get_prelim_stats(team, tournament.prelim_rounds)
        ranked.append((team, ballots, pd))
    ranked.sort(key=lambda item: (item[1], item[2]), reverse=True)
    return ranked


def get_round_winner(round_obj):
    ballots = [ballot for ballot in counted_ballots_for_round(round_obj) if ballot.submit]
    if len(ballots) < round_obj.pairing.ballots_counted():
        return None
    p_ballots = sum(ballot.p_ballot for ballot in ballots)
    d_ballots = sum(ballot.d_ballot for ballot in ballots)
    p_pd = sum(ballot.p_pd for ballot in ballots)
    d_pd = sum(ballot.d_pd for ballot in ballots)
    if p_ballots > d_ballots:
        return round_obj.p_team
    if d_ballots > p_ballots:
        return round_obj.d_team
    if p_pd > d_pd:
        return round_obj.p_team
    if d_pd > p_pd:
        return round_obj.d_team
    return None


def get_elim_teams_for_round(tournament, round_num):
    if not tournament.is_elim_round(round_num):
        return []
    if round_num == tournament.prelim_rounds + 1:
        ranked = rank_teams_for_break(tournament)
        return [team for team, ballots, pd in ranked[:tournament.elim_break_size]]

    previous_pairing = Pairing.objects.filter(tournament=tournament, round_num=round_num - 1).first()
    if not previous_pairing:
        return []
    winners = [winner for winner in (get_round_winner(round_obj) for round_obj in previous_pairing.rounds.all()) if winner]
    ranked_winners = rank_teams_for_break(tournament, winners)
    return [team for team, ballots, pd in ranked_winners]


def build_elim_pairings(tournament, round_num):
    teams = get_elim_teams_for_round(tournament, round_num)
    if not teams:
        return []
    teams = teams[: tournament.elim_break_size] if round_num == tournament.prelim_rounds + 1 else teams
    pairings = []
    while len(teams) >= 2:
        pairings.append((teams.pop(0), teams.pop(-1)))
    return pairings


def lock_prelim_results(tournament):
    teams = sort_teams(Team.objects.filter(user__tournament=tournament))
    if teams and all(team.prelim_seed for team in teams):
        return
    for index, team in enumerate(teams, start=1):
        team.prelim_seed = index
        team.locked_prelim_ballots = team.total_ballots
        team.locked_prelim_cs = team.total_cs
        team.locked_prelim_pd = team.total_pd
        team.save(update_fields=['prelim_seed', 'locked_prelim_ballots', 'locked_prelim_cs', 'locked_prelim_pd'])


def get_pairing_capacity(tournament, round_num):
    if tournament.is_elim_round(round_num):
        return max(1, len(build_elim_pairings(tournament, round_num)))
    if tournament.split_division:
        division_counts = [
            Team.objects.filter(user__tournament=tournament, division=division).count()
            for division in ['Disney', 'Universal']
        ]
        team_count = max(division_counts) if division_counts else 0
    else:
        team_count = Team.objects.filter(user__tournament=tournament).count()
    return max(1, (team_count + 1) // 2)


def get_round_title(tournament, round_num):
    return tournament.get_round_label(round_num)


def build_prelim_pairings(tournament, round_num, division=None):
    if tournament.is_elim_round(round_num):
        return []
    queryset = Team.objects.filter(user__tournament=tournament)
    if division:
        queryset = queryset.filter(division=division)
    teams = [
        team for team in sort_teams(list(queryset))
        if team.current_rounds() < tournament.team_prelim_rounds
    ]
    if not teams:
        return []
    if round_num % 2 == 0:
        d_teams = [team for team in teams if team.next_side(round_num) == 'd']
        p_teams = [team for team in teams if team.next_side(round_num) == 'p']
        return list(zip(p_teams, d_teams))

    p_teams = []
    d_teams = []
    for i in range(0, len(teams), 2):
        if i + 1 >= len(teams):
            break
        if random.randint(0, 1):
            p_teams.append(teams[i])
            d_teams.append(teams[i + 1])
        else:
            p_teams.append(teams[i + 1])
            d_teams.append(teams[i])
    return list(zip(p_teams, d_teams))


def get_group_teams_for_pairing(tournament, division=None):
    queryset = Team.objects.filter(user__tournament=tournament)
    if division:
        queryset = queryset.filter(division=division)
    teams = list(queryset.order_by('team_name'))
    return teams


def get_side_targets(teams, prelim_rounds):
    return get_side_targets_for_appearances({team.pk: prelim_rounds for team in teams}, teams)


def get_side_targets_for_appearances(appearance_targets, teams):
    base_target = {
        team.pk: appearance_targets[team.pk] // 2
        for team in teams
    }
    extra_p_slots = sum(appearance_targets.values()) // 2 - sum(base_target.values())
    shuffled = teams[:]
    random.shuffle(shuffled)
    for team in shuffled[:extra_p_slots]:
        base_target[team.pk] += 1
    return base_target


def choose_petitioner_side(teams, side_targets, side_counts, round_num, prelim_rounds):
    petitioner_slots = len(teams) // 2
    remaining_rounds = prelim_rounds - round_num + 1
    must_petitioner = [
        team for team in teams
        if side_targets[team.pk] - side_counts[team.pk] >= remaining_rounds
    ]
    if len(must_petitioner) > petitioner_slots:
        return None
    eligible = [team for team in teams if side_counts[team.pk] < side_targets[team.pk] and team not in must_petitioner]
    random.shuffle(eligible)
    petitioner_teams = must_petitioner[:]
    petitioner_teams.extend(eligible[: petitioner_slots - len(must_petitioner)])
    if len(petitioner_teams) != petitioner_slots:
        return None
    respondents = [team for team in teams if team not in petitioner_teams]
    for team in respondents:
        if side_targets[team.pk] - side_counts[team.pk] > remaining_rounds - 1:
            return None
    return petitioner_teams, respondents


def schedule_counts(schedule):
    appearance_counts = defaultdict(int)
    side_counts = defaultdict(int)
    for pairs in schedule:
        for p_team, d_team in pairs:
            appearance_counts[p_team.pk] += 1
            appearance_counts[d_team.pk] += 1
            side_counts[p_team.pk] += 1
    return appearance_counts, side_counts


def validate_prelim_schedule(schedule, teams, appearance_targets, side_targets):
    appearance_counts, side_counts = schedule_counts(schedule)
    errors = []
    for team in teams:
        appearances = appearance_counts[team.pk]
        petitioner_rounds = side_counts[team.pk]
        if appearances != appearance_targets[team.pk]:
            errors.append(
                f'{team} is scheduled for {appearances} preliminary round(s), '
                f'but should be scheduled for {appearance_targets[team.pk]}.'
            )
        if petitioner_rounds != side_targets[team.pk]:
            errors.append(
                f'{team} is petitioner in {petitioner_rounds} preliminary round(s), '
                f'but should be petitioner in {side_targets[team.pk]}.'
            )
        if appearance_targets[team.pk] > 1 and petitioner_rounds in [0, appearance_targets[team.pk]]:
            errors.append(f'{team} is scheduled on only one side.')
    return errors


def choose_petitioner_side_for_schedule(
    playing_teams,
    all_teams,
    appearance_targets,
    appearance_counts,
    side_targets,
    side_counts,
    round_num,
    prelim_rounds,
):
    for _ in range(100):
        side_split = choose_petitioner_side(playing_teams, side_targets, side_counts, round_num, prelim_rounds)
        if not side_split:
            return None
        petitioner_teams, respondent_teams = side_split
        petitioner_ids = {team.pk for team in petitioner_teams}
        playing_ids = {team.pk for team in playing_teams}
        feasible = True

        for team in all_teams:
            remaining_appearances = appearance_targets[team.pk] - appearance_counts[team.pk] - (1 if team.pk in playing_ids else 0)
            remaining_petitioner = side_targets[team.pk] - side_counts[team.pk] - (1 if team.pk in petitioner_ids else 0)
            if remaining_appearances < 0 or remaining_petitioner < 0 or remaining_petitioner > remaining_appearances:
                feasible = False
                break
        if feasible:
            return petitioner_teams, respondent_teams
    return None


def pairing_penalty(p_team, d_team, seen_matchups, seen_side_matchups):
    penalty = 0
    if (p_team.pk, d_team.pk) in seen_side_matchups:
        penalty += 10000
    if p_team.school and d_team.school and p_team.school == d_team.school:
        penalty += 1000
    if frozenset((p_team.pk, d_team.pk)) in seen_matchups:
        penalty += 100
    return penalty


def get_pairing_conflict_reasons(p_team, d_team, seen_matchups, seen_side_matchups):
    reasons = []
    if (p_team.pk, d_team.pk) in seen_side_matchups:
        reasons.append('Repeat matchup on same sides')
    if p_team.school and d_team.school and p_team.school == d_team.school:
        reasons.append(f'Same school: {p_team.school}')
    if frozenset((p_team.pk, d_team.pk)) in seen_matchups:
        reasons.append('Repeat matchup')
    return reasons


def round_pairing_penalty(pairs, seen_matchups, seen_side_matchups):
    return sum(pairing_penalty(p_team, d_team, seen_matchups, seen_side_matchups) for p_team, d_team in pairs)


def improve_round_pairs(pairs, seen_matchups, seen_side_matchups):
    pairs = pairs[:]
    best_penalty = round_pairing_penalty(pairs, seen_matchups, seen_side_matchups)
    improved = True
    while improved:
        improved = False
        for i in range(len(pairs)):
            for j in range(i + 1, len(pairs)):
                swapped = pairs[:]
                swapped[i] = (pairs[i][0], pairs[j][1])
                swapped[j] = (pairs[j][0], pairs[i][1])
                swapped_penalty = round_pairing_penalty(swapped, seen_matchups, seen_side_matchups)
                if swapped_penalty < best_penalty:
                    pairs = swapped
                    best_penalty = swapped_penalty
                    improved = True
                    break
            if improved:
                break
    return pairs, best_penalty


def match_round_teams(p_teams, d_teams, seen_matchups, seen_side_matchups, attempts=250):
    best_pairs = None
    best_penalty = None
    p_teams = p_teams[:]
    d_teams = d_teams[:]
    for _ in range(attempts):
        random.shuffle(p_teams)
        random.shuffle(d_teams)
        pairs = list(zip(p_teams, d_teams))
        pairs, penalty = improve_round_pairs(pairs, seen_matchups, seen_side_matchups)
        if best_penalty is None or penalty < best_penalty:
            best_pairs = pairs[:]
            best_penalty = penalty
        if penalty == 0:
            break
    return best_pairs, best_penalty


def get_partial_round_pair_counts(team_count, team_rounds, prelim_rounds, distribution='spread'):
    max_pairs = team_count // 2
    total_slots = team_count * team_rounds
    if total_slots % 2 != 0 or max_pairs < 1:
        return None
    total_pairs = total_slots // 2
    if team_rounds > prelim_rounds or total_pairs < prelim_rounds or total_pairs > prelim_rounds * max_pairs:
        return None
    pair_counts = [1] * prelim_rounds
    remaining = total_pairs - prelim_rounds
    if distribution == 'concentrate':
        for index in range(prelim_rounds):
            add = min(max_pairs - pair_counts[index], remaining)
            pair_counts[index] += add
            remaining -= add
            if remaining == 0:
                break
    else:
        index = 0
        while remaining > 0:
            if pair_counts[index] < max_pairs:
                pair_counts[index] += 1
                remaining -= 1
            index = (index + 1) % prelim_rounds
    if remaining != 0:
        return None
    return pair_counts


def build_partial_prelim_schedule(teams, prelim_rounds, team_rounds, distribution='spread'):
    if len(teams) < 2:
        return None, ['Not enough teams to generate preliminary rounds.'], []
    pair_counts = get_partial_round_pair_counts(len(teams), team_rounds, prelim_rounds, distribution)
    if not pair_counts:
        return None, ['Auto-generation settings are not feasible for this team count and number of preliminary rounds.'], []

    appearance_targets = {team.pk: team_rounds for team in teams}
    for _ in range(300):
        appearance_counts = defaultdict(int)
        side_counts = defaultdict(int)
        side_targets = get_side_targets_for_appearances(appearance_targets, teams)
        seen_matchups = set()
        seen_side_matchups = set()
        schedule = []
        conflicts = []
        failed = False
        for round_num, pair_count in enumerate(pair_counts, start=1):
            playing_teams = choose_playing_teams_for_byebuster(
                teams, appearance_targets, appearance_counts, pair_count * 2, round_num, prelim_rounds
            )
            if not playing_teams:
                failed = True
                break
            side_split = choose_petitioner_side_for_schedule(
                playing_teams,
                teams,
                appearance_targets,
                appearance_counts,
                side_targets,
                side_counts,
                round_num,
                prelim_rounds,
            )
            if not side_split:
                failed = True
                break
            petitioner_teams, respondent_teams = side_split
            pairs, penalty = match_round_teams(petitioner_teams, respondent_teams, seen_matchups, seen_side_matchups)
            if not pairs:
                failed = True
                break
            schedule.append(pairs)
            for p_team, d_team in pairs:
                conflict_reasons = get_pairing_conflict_reasons(p_team, d_team, seen_matchups, seen_side_matchups)
                if conflict_reasons:
                    conflicts.append({
                        'round': round_num,
                        'p_team': str(p_team),
                        'd_team': str(d_team),
                        'reasons': conflict_reasons,
                    })
                for team in [p_team, d_team]:
                    appearance_counts[team.pk] += 1
                side_counts[p_team.pk] += 1
                seen_matchups.add(frozenset((p_team.pk, d_team.pk)))
                seen_side_matchups.add((p_team.pk, d_team.pk))
        validation_errors = [] if failed else validate_prelim_schedule(schedule, teams, appearance_targets, side_targets)
        if not failed and not validation_errors:
            return schedule, [], conflicts
    return None, ['Auto-generation could not resolve round-count and side-balance constraints. Please edit pairings manually.'], []


def build_random_prelim_schedule(teams, prelim_rounds, team_rounds=None):
    team_rounds = team_rounds or prelim_rounds
    if team_rounds < prelim_rounds:
        return build_partial_prelim_schedule(teams, prelim_rounds, team_rounds)
    if len(teams) % 2 != 0:
        return None, ['Auto-generation requires an even number of teams in each pairing pool.']
    if len(teams) < 2:
        return None, ['Not enough teams to generate preliminary rounds.']

    for _ in range(300):
        appearance_targets = {team.pk: prelim_rounds for team in teams}
        appearance_counts = defaultdict(int)
        side_targets = get_side_targets(teams, prelim_rounds)
        side_counts = defaultdict(int)
        seen_matchups = set()
        seen_side_matchups = set()
        schedule = []
        conflicts = []
        failed = False
        for round_num in range(1, prelim_rounds + 1):
            side_split = choose_petitioner_side(teams, side_targets, side_counts, round_num, prelim_rounds)
            if not side_split:
                failed = True
                break
            petitioner_teams, respondent_teams = side_split
            pairs, penalty = match_round_teams(petitioner_teams, respondent_teams, seen_matchups, seen_side_matchups)
            if not pairs:
                failed = True
                break
            schedule.append(pairs)
            for p_team, d_team in pairs:
                conflict_reasons = get_pairing_conflict_reasons(p_team, d_team, seen_matchups, seen_side_matchups)
                if conflict_reasons:
                    conflicts.append({
                        'round': round_num,
                        'p_team': str(p_team),
                        'd_team': str(d_team),
                        'reasons': conflict_reasons,
                    })
                side_counts[p_team.pk] += 1
                appearance_counts[p_team.pk] += 1
                appearance_counts[d_team.pk] += 1
                seen_matchups.add(frozenset((p_team.pk, d_team.pk)))
                seen_side_matchups.add((p_team.pk, d_team.pk))
        validation_errors = [] if failed else validate_prelim_schedule(schedule, teams, appearance_targets, side_targets)
        if not failed and not validation_errors:
            return schedule, [], conflicts
    return None, ['Auto-generation could not resolve school/rematch constraints. Please edit pairings manually.'], []


def get_byebuster_distribution_order(prelim_rounds, lightest_round=None):
    if not lightest_round:
        return list(range(prelim_rounds))
    lightest_index = lightest_round - 1
    if lightest_index < 0 or lightest_index >= prelim_rounds:
        return list(range(prelim_rounds))
    return list(range(lightest_index + 1, prelim_rounds)) + list(range(0, lightest_index)) + [lightest_index]


def get_byebuster_round_pair_counts(team_count, counted_rounds, prelim_rounds, distribution, lightest_round=None):
    max_pairs = (team_count - 1) // 2
    total_pairs = (team_count * counted_rounds + 1) // 2
    if counted_rounds >= prelim_rounds or total_pairs < prelim_rounds or total_pairs > prelim_rounds * max_pairs:
        return None
    pair_counts = [1] * prelim_rounds
    remaining = total_pairs - prelim_rounds
    distribution_order = get_byebuster_distribution_order(prelim_rounds, lightest_round)
    if distribution == 'concentrate':
        for index in distribution_order:
            add = min(max_pairs - pair_counts[index], remaining)
            pair_counts[index] += add
            remaining -= add
            if remaining == 0:
                break
    else:
        index = 0
        while remaining > 0:
            round_index = distribution_order[index]
            if pair_counts[round_index] < max_pairs:
                pair_counts[round_index] += 1
                remaining -= 1
            index = (index + 1) % len(distribution_order)
    if remaining != 0:
        return None
    return pair_counts


def choose_byebuster_team(teams, school_choice):
    if school_choice == 'random':
        return random.choice(teams)
    candidates = [team for team in teams if team.school == school_choice]
    if not candidates:
        return None
    return random.choice(candidates)


def choose_playing_teams_for_byebuster(teams, appearance_targets, appearance_counts, playing_count, round_num, prelim_rounds):
    remaining_rounds = prelim_rounds - round_num + 1
    must_play = [
        team for team in teams
        if appearance_targets[team.pk] - appearance_counts[team.pk] >= remaining_rounds
    ]
    if len(must_play) > playing_count:
        return None
    eligible = [
        team for team in teams
        if appearance_counts[team.pk] < appearance_targets[team.pk] and team not in must_play
    ]
    eligible.sort(key=lambda team: appearance_targets[team.pk] - appearance_counts[team.pk], reverse=True)
    tied_groups = defaultdict(list)
    for team in eligible:
        tied_groups[appearance_targets[team.pk] - appearance_counts[team.pk]].append(team)
    eligible = []
    for key in sorted(tied_groups.keys(), reverse=True):
        group = tied_groups[key]
        random.shuffle(group)
        eligible.extend(group)
    playing = must_play + eligible[: playing_count - len(must_play)]
    if len(playing) != playing_count:
        return None
    for team in teams:
        remaining_after = appearance_targets[team.pk] - appearance_counts[team.pk] - (1 if team in playing else 0)
        if remaining_after > remaining_rounds - 1:
            return None
    return playing


def build_byebuster_prelim_schedule(teams, prelim_rounds, counted_rounds, school_choice, distribution, lightest_round=None):
    if len(teams) % 2 != 1:
        return None, ['Byebuster generation only applies to odd team pools.'], [], None
    byebuster_team = choose_byebuster_team(teams, school_choice)
    if not byebuster_team:
        return None, [f'No team found from school "{school_choice}" in an odd pairing pool.'], [], None
    pair_counts = get_byebuster_round_pair_counts(len(teams), counted_rounds, prelim_rounds, distribution, lightest_round)
    if not pair_counts:
        return None, ['Byebuster settings are not feasible for this team count and number of preliminary rounds.'], [], None

    appearance_targets = {team.pk: counted_rounds for team in teams}
    appearance_targets[byebuster_team.pk] = counted_rounds + 1

    for _ in range(300):
        appearance_counts = defaultdict(int)
        side_counts = defaultdict(int)
        side_targets = get_side_targets_for_appearances(appearance_targets, teams)
        seen_matchups = set()
        seen_side_matchups = set()
        schedule = []
        conflicts = []
        failed = False
        for round_num, pair_count in enumerate(pair_counts, start=1):
            playing_count = pair_count * 2
            playing_teams = choose_playing_teams_for_byebuster(
                teams, appearance_targets, appearance_counts, playing_count, round_num, prelim_rounds
            )
            if not playing_teams:
                failed = True
                break
            side_split = choose_petitioner_side_for_schedule(
                playing_teams,
                teams,
                appearance_targets,
                appearance_counts,
                side_targets,
                side_counts,
                round_num,
                prelim_rounds,
            )
            if not side_split:
                failed = True
                break
            petitioner_teams, respondent_teams = side_split
            pairs, penalty = match_round_teams(petitioner_teams, respondent_teams, seen_matchups, seen_side_matchups)
            if not pairs:
                failed = True
                break
            schedule.append(pairs)
            for p_team, d_team in pairs:
                conflict_reasons = get_pairing_conflict_reasons(p_team, d_team, seen_matchups, seen_side_matchups)
                if conflict_reasons:
                    conflicts.append({
                        'round': round_num,
                        'p_team': str(p_team),
                        'd_team': str(d_team),
                        'reasons': conflict_reasons,
                    })
                for team in [p_team, d_team]:
                    appearance_counts[team.pk] += 1
                side_counts[p_team.pk] += 1
                seen_matchups.add(frozenset((p_team.pk, d_team.pk)))
                seen_side_matchups.add((p_team.pk, d_team.pk))
        validation_errors = [] if failed else validate_prelim_schedule(schedule, teams, appearance_targets, side_targets)
        if not failed and not validation_errors:
            return schedule, [], conflicts, byebuster_team
    return None, ['Auto-generation could not resolve byebuster round-count and side-balance constraints.'], [], None


def judge_preside_rank(judge):
    if judge.preside == 1:
        return 2
    if judge.preside == 2:
        return 1
    return 0


def judge_available_round_count(judge, tournament):
    total_rounds = min(max(tournament.total_rounds, 1), 9)
    return sum(1 for round_num in range(1, total_rounds + 1) if judge.get_availability(round_num))


def judge_assignment_sort_key(judge, tournament, prefer_presiding=True):
    preside_rank = -judge_preside_rank(judge) if prefer_presiding else judge_preside_rank(judge)
    return (
        judge_available_round_count(judge, tournament),
        preside_rank,
        judge.user.username,
    )


def judge_can_cover_round(judge, round_obj, round_num, used_judges):
    if judge in used_judges:
        return False
    if not judge.get_availability(round_num):
        return False
    for team in round_obj.teams:
        if team in judge.conflicts.all():
            return False
    p_judged, d_judged = judge.judged(round_num)
    if round_obj.p_team in p_judged:
        return False
    if round_obj.d_team in d_judged:
        return False
    if round_obj.pairing.tournament.conflict_other_side:
        if round_obj.p_team in d_judged or round_obj.d_team in p_judged:
            return False
    return True


def round_assignment_sort_key(round_obj, round_num, judge_pool):
    unused_judges = set()
    valid_presiding_count = sum(
        1 for judge in judge_pool
        if judge.preside > 0 and judge_can_cover_round(judge, round_obj, round_num, unused_judges)
    )
    valid_judge_count = sum(
        1 for judge in judge_pool
        if judge_can_cover_round(judge, round_obj, round_num, unused_judges)
    )
    return (
        valid_presiding_count,
        valid_judge_count,
        round_obj.pairing.division or '',
        round_obj.courtroom or '',
    )


def judge_preserves_existing_assignments(judge, round_obj):
    tournament = round_obj.pairing.tournament
    for existing_round in judge.rounds:
        if existing_round == round_obj:
            continue
        if existing_round.pairing.tournament != tournament:
            continue
        if round_obj.p_team == existing_round.p_team or round_obj.d_team == existing_round.d_team:
            return False
        if tournament.conflict_other_side:
            if round_obj.p_team == existing_round.d_team or round_obj.d_team == existing_round.p_team:
                return False
    return True


def judge_can_cover_scoring_round(judge, round_obj, round_num, used_judges, waive_seen_conflicts=False):
    if judge in used_judges:
        return False
    if not judge.get_availability(round_num):
        return False
    for team in round_obj.teams:
        if team in judge.conflicts.all():
            return False
    p_judged, d_judged = judge.judged(round_num)
    if round_obj.p_team in p_judged:
        return False
    if round_obj.d_team in d_judged:
        return False
    if not waive_seen_conflicts and round_obj.pairing.tournament.conflict_other_side:
        if round_obj.p_team in d_judged or round_obj.d_team in p_judged:
            return False
    return True


def judge_preserves_scoring_assignments(judge, round_obj, waive_seen_conflicts=False):
    tournament = round_obj.pairing.tournament
    for existing_round in judge.rounds:
        if existing_round == round_obj:
            continue
        if existing_round.pairing.tournament != tournament:
            continue
        if round_obj.p_team == existing_round.p_team or round_obj.d_team == existing_round.d_team:
            return False
        if not waive_seen_conflicts and tournament.conflict_other_side:
            if round_obj.p_team == existing_round.d_team or round_obj.d_team == existing_round.p_team:
                return False
    return True


def assign_judges_for_rounds(tournament, round_num, round_objects, target_judges=None):
    if target_judges is None:
        target_judges = max(tournament.judges, tournament.required_judges)
    target_judges = min(max(1, target_judges), tournament.get_max_judges_for_round(round_num))
    available = list(Judge.objects.filter(user__tournament=tournament))
    checked_in = [judge for judge in available if judge.checkin and judge.get_availability(round_num)]
    judge_pool = checked_in if checked_in else [judge for judge in available if judge.get_availability(round_num)]
    judge_pool = sorted(judge_pool, key=lambda judge: judge_assignment_sort_key(judge, tournament))
    round_objects = sorted(round_objects, key=lambda round_obj: round_assignment_sort_key(round_obj, round_num, judge_pool))
    used_judges = set()

    for round_obj in round_objects:
        valid_presiding = [judge for judge in judge_pool if judge_can_cover_round(judge, round_obj, round_num, used_judges) and judge.preside > 0]
        if not valid_presiding:
            return False
        round_obj.presiding_judge = valid_presiding[0]
        used_judges.add(valid_presiding[0])

        scoring_candidates = [
            judge for judge in sorted(judge_pool, key=lambda judge: judge_assignment_sort_key(judge, tournament, prefer_presiding=False))
            if judge_can_cover_round(judge, round_obj, round_num, used_judges)
        ]
        needed_scoring = max(0, target_judges - 1)
        if len(scoring_candidates) < needed_scoring:
            return False
        selected = scoring_candidates[:needed_scoring]
        round_obj.scoring_judge = selected[0] if needed_scoring >= 1 else None
        round_obj.extra_judge = selected[1] if needed_scoring >= 2 else None
        for judge in selected:
            used_judges.add(judge)
    return True


def assign_presiding_judges_for_existing_round(tournament, round_num):
    pairings = Pairing.objects.filter(tournament=tournament, round_num=round_num)
    round_objects = list(Round.objects.filter(pairing__in=pairings).order_by('pairing__division', 'courtroom'))
    if not round_objects:
        return False

    available = list(Judge.objects.filter(user__tournament=tournament))
    checked_in = [judge for judge in available if judge.checkin and judge.get_availability(round_num)]
    judge_pool = checked_in if checked_in else [judge for judge in available if judge.get_availability(round_num)]
    judge_pool = sorted(judge_pool, key=lambda judge: judge_assignment_sort_key(judge, tournament))
    round_objects = sorted(round_objects, key=lambda round_obj: round_assignment_sort_key(round_obj, round_num, judge_pool))
    used_judges = {
        judge
        for round_obj in round_objects
        for judge in [round_obj.scoring_judge, round_obj.extra_judge, *list(round_obj.additional_judges.all())]
        if judge
    }

    for round_obj in round_objects:
        valid_presiding = [
            judge for judge in judge_pool
            if judge_can_cover_round(judge, round_obj, round_num, used_judges) and judge.preside > 0
        ]
        if not valid_presiding:
            return False
        round_obj.presiding_judge = valid_presiding[0]
        used_judges.add(valid_presiding[0])

    for round_obj in round_objects:
        round_obj.save(update_fields=['presiding_judge'])
    for pairing in pairings:
        pairing.final_submit = True
        pairing.save(update_fields=['final_submit'])
        sync_ballots_for_pairing(pairing)
    return True


@user_passes_test(lambda u: u.is_staff)
def assign_judges(request, round_num):
    tournament = request.user.tournament
    if tournament.is_elim_round(round_num):
        set_pairing_banner(request, tournament, ['Automatic judge assignment is only available for preliminary rounds.'])
        return redirect('tourney:pairing_index')

    if assign_presiding_judges_for_existing_round(tournament, round_num):
        set_pairing_banner(request, tournament, [f'Assigned presiding judges for {tournament.get_round_label(round_num)}.'])
    else:
        set_pairing_banner(
            request,
            tournament,
            [f'Unable to assign presiding judges for {tournament.get_round_label(round_num)}. Check judge availability, check-in status, and conflicts.'],
        )
    return redirect('tourney:pairing_index')


def assign_free_scoring_judges_for_round(tournament, round_num):
    pairings = Pairing.objects.filter(tournament=tournament, round_num=round_num)
    round_objects = list(Round.objects.filter(pairing__in=pairings).order_by('pairing__division', 'courtroom'))
    if not round_objects:
        return 0

    judge_pool = list(Judge.objects.filter(user__tournament=tournament))
    judge_pool = [judge for judge in judge_pool if judge.get_availability(round_num)]
    judge_pool = sorted(judge_pool, key=lambda judge: (
        -judge.checkin,
        *judge_assignment_sort_key(judge, tournament, prefer_presiding=False),
    ))
    used_judges = {
        judge
        for round_obj in round_objects
        for judge in [round_obj.presiding_judge, round_obj.scoring_judge, round_obj.extra_judge, *list(round_obj.additional_judges.all())]
        if judge
    }
    assigned_count = 0

    changed_rounds = {}
    for field_name in ['scoring_judge', 'extra_judge']:
        for round_obj in round_objects:
            if getattr(round_obj, field_name):
                continue
            candidates = [
                judge for judge in judge_pool
                if judge_can_cover_round(judge, round_obj, round_num, used_judges)
                and judge_preserves_existing_assignments(judge, round_obj)
            ]
            if not candidates:
                continue
            judge = candidates[0]
            setattr(round_obj, field_name, judge)
            used_judges.add(judge)
            changed_rounds.setdefault(round_obj, set()).add(field_name)
            assigned_count += 1

    for round_obj, changed_fields in changed_rounds.items():
        round_obj.save(update_fields=sorted(changed_fields))

    for pairing in pairings:
        sync_ballots_for_pairing(pairing)
    return assigned_count


def scoring_judge_sort_key(judge, tournament):
    return (
        -judge.checkin,
        *judge_assignment_sort_key(judge, tournament, prefer_presiding=False),
    )


def scoring_slot_field_count(round_obj, field_name):
    if field_name == 'scoring_judge':
        return 1
    if field_name == 'extra_judge':
        return 2
    return 3


def round_scoring_judges(round_obj):
    return [judge for judge in [round_obj.scoring_judge, round_obj.extra_judge] if judge]


def global_scoring_slot_sort_key(
    slot,
    priority_presiding_judge_ids,
    priority_presiding_load,
    scoring_slots_by_round,
    candidate_count,
):
    round_obj, field_name = slot
    round_num = round_obj.pairing.round_num
    priority_rank = round_obj.presiding_judge_id in priority_presiding_judge_ids
    priority_load = priority_presiding_load.get(round_obj.presiding_judge_id, 0) if priority_rank else 0
    return (
        not priority_rank,
        priority_load,
        candidate_count,
        scoring_slots_by_round[round_num],
        scoring_slot_field_count(round_obj, field_name),
        round_num,
        round_obj.pairing.division or '',
        round_obj.courtroom or '',
    )


def global_scoring_judge_sort_key(judge, tournament, round_num, scoring_round_loads, total_loads):
    return (
        scoring_round_loads[(judge.pk, round_num)],
        total_loads[judge.pk],
        judge_available_round_count(judge, tournament),
        -judge.checkin,
        judge.user.username,
    )


def assign_global_scoring_judges(tournament, priority_judges=None, waive_seen_conflicts=False):
    priority_presiding_judge_ids = {judge.pk for judge in (priority_judges or [])}
    pairings = Pairing.objects.filter(tournament=tournament, round_num__lte=tournament.prelim_rounds)
    round_objects = list(
        Round.objects.filter(pairing__in=pairings).select_related(
            'pairing',
            'pairing__tournament',
            'p_team',
            'd_team',
        ).order_by('pairing__round_num', 'pairing__division', 'courtroom')
    )
    if not round_objects:
        return 0

    judge_pool = list(Judge.objects.filter(user__tournament=tournament).select_related('user'))
    judge_pool = sorted(judge_pool, key=lambda judge: scoring_judge_sort_key(judge, tournament))
    used_judges_by_round = defaultdict(set)
    scoring_round_loads = defaultdict(int)
    total_scoring_loads = defaultdict(int)
    priority_presiding_load = defaultdict(int)
    scoring_slots_by_round = defaultdict(int)
    for round_obj in round_objects:
        for judge in [round_obj.presiding_judge, round_obj.scoring_judge, round_obj.extra_judge, *list(round_obj.additional_judges.all())]:
            if judge:
                used_judges_by_round[round_obj.pairing.round_num].add(judge)
        for judge in round_scoring_judges(round_obj):
            scoring_round_loads[(judge.pk, round_obj.pairing.round_num)] += 1
            total_scoring_loads[judge.pk] += 1
            scoring_slots_by_round[round_obj.pairing.round_num] += 1
            if round_obj.presiding_judge_id in priority_presiding_judge_ids:
                priority_presiding_load[round_obj.presiding_judge_id] += 1

    assigned_count = 0
    for field_name in ['scoring_judge', 'extra_judge']:
        while True:
            slots = []
            candidates_by_slot = {}
            for round_obj in round_objects:
                if getattr(round_obj, field_name):
                    continue
                if field_name == 'extra_judge' and not round_obj.scoring_judge:
                    continue
                round_num = round_obj.pairing.round_num
                candidates = [
                    judge for judge in judge_pool
                    if judge_can_cover_scoring_round(
                        judge,
                        round_obj,
                        round_num,
                        used_judges_by_round[round_num],
                        waive_seen_conflicts=waive_seen_conflicts,
                    )
                    and judge_preserves_scoring_assignments(
                        judge,
                        round_obj,
                        waive_seen_conflicts=waive_seen_conflicts,
                    )
                ]
                if candidates:
                    slot = (round_obj, field_name)
                    slots.append(slot)
                    candidates_by_slot[slot] = candidates
            if not slots:
                break
            slot = sorted(
                slots,
                key=lambda candidate_slot: global_scoring_slot_sort_key(
                    candidate_slot,
                    priority_presiding_judge_ids,
                    priority_presiding_load,
                    scoring_slots_by_round,
                    len(candidates_by_slot[candidate_slot]),
                ),
            )[0]
            round_obj, field_name = slot
            round_num = round_obj.pairing.round_num
            candidates = sorted(
                candidates_by_slot[slot],
                key=lambda judge: global_scoring_judge_sort_key(
                    judge,
                    tournament,
                    round_num,
                    scoring_round_loads,
                    total_scoring_loads,
                ),
            )
            judge = candidates[0]
            setattr(round_obj, field_name, judge)
            round_obj.save(update_fields=[field_name])
            used_judges_by_round[round_num].add(judge)
            scoring_round_loads[(judge.pk, round_num)] += 1
            total_scoring_loads[judge.pk] += 1
            scoring_slots_by_round[round_num] += 1
            if round_obj.presiding_judge_id in priority_presiding_judge_ids:
                priority_presiding_load[round_obj.presiding_judge_id] += 1
            assigned_count += 1

    for pairing in pairings:
        sync_ballots_for_pairing(pairing)
    return assigned_count


@user_passes_test(lambda u: u.is_staff)
def assign_global_scoring_judges_view(request):
    tournament = request.user.tournament
    if request.method == 'POST':
        form = GlobalScoringJudgeForm(request.POST, tournament=tournament)
        if form.is_valid():
            priority_judges = []
            if request.POST.get('mode') == 'priority':
                priority_judges = form.cleaned_data.get('priority_judges')
            assigned_count = assign_global_scoring_judges(
                tournament,
                priority_judges=priority_judges,
                waive_seen_conflicts=form.cleaned_data.get('waive_seen_conflicts'),
            )
            set_pairing_banner(
                request,
                tournament,
                [f'Assigned {assigned_count} scoring judge slot(s) across preliminary rounds.'],
            )
            return redirect('tourney:pairing_index')
    else:
        form = GlobalScoringJudgeForm(tournament=tournament)
    return render(request, 'tourney/pairing/assign_global_scoring_judges.html', {'form': form})


@user_passes_test(lambda u: u.is_staff)
def assign_free_scoring_judges(request, round_num):
    tournament = request.user.tournament
    if tournament.is_elim_round(round_num):
        set_pairing_banner(request, tournament, ['Free scoring judge assignment is only available for preliminary rounds.'])
        return redirect('tourney:pairing_index')

    assigned_count = assign_free_scoring_judges_for_round(tournament, round_num)
    extra = request.session.get('extra', {})
    if extra.get('tournament_id') != tournament.pk:
        extra = {}
    extra['errors'] = [
        f'Assigned {assigned_count} free available judge(s) as scoring judges for {tournament.get_round_label(round_num)}. '
        f'Only {tournament.judges} ballot(s) per round will count toward results.'
    ]
    extra['tournament_id'] = tournament.pk
    request.session['extra'] = extra
    return redirect('tourney:pairing_index')


@user_passes_test(lambda u: u.is_staff)
def clear_scoring_judges(request, round_num):
    tournament = request.user.tournament
    pairings = Pairing.objects.filter(tournament=tournament, round_num=round_num)
    cleared_count = 0
    for round_obj in Round.objects.filter(pairing__in=pairings):
        cleared_count += len([judge for judge in [round_obj.scoring_judge, round_obj.extra_judge] if judge])
        cleared_count += round_obj.additional_judges.count()
        round_obj.scoring_judge = None
        round_obj.extra_judge = None
        round_obj.save(update_fields=['scoring_judge', 'extra_judge'])
        round_obj.additional_judges.clear()
    for pairing in pairings:
        sync_ballots_for_pairing(pairing)
    set_pairing_banner(
        request,
        tournament,
        [f'Cleared {cleared_count} scoring judge slot(s) for {tournament.get_round_label(round_num)}.'],
    )
    return redirect('tourney:pairing_index')


def get_pairing_letters(division, count):
    start_index = 8 if division == 'Universal' else 0
    letters = []
    for offset in range(count):
        number = start_index + offset
        label = ''
        while True:
            number, remainder = divmod(number, 26)
            label = string.ascii_uppercase[remainder] + label
            if number == 0:
                break
            number -= 1
        letters.append(label)
    return letters


@user_passes_test(lambda u: u.is_staff)
def generate_prelim_pairings(request):
    tournament = request.user.tournament
    if not tournament.randomize_prelims:
        set_pairing_banner(request, tournament, ['Random preliminary generation is turned off in tournament settings.'])
        return redirect('tourney:pairing_index')
    existing_pairings = Pairing.objects.filter(tournament=tournament)
    if existing_pairings.filter(round_num__gt=tournament.prelim_rounds).exists():
        set_pairing_banner(request, tournament, ['Delete existing pairings before generating all preliminary rounds automatically.'])
        return redirect('tourney:pairing_index')

    divisions = ['Disney', 'Universal'] if tournament.split_division else [None]
    byebuster_needed = any(
        len(get_group_teams_for_pairing(tournament, division)) * tournament.team_prelim_rounds % 2 == 1
        for division in divisions
    )
    byebuster_options = None
    if byebuster_needed:
        if request.method != 'POST':
            form = ByebusterGenerateForm(tournament=tournament, divisions=divisions)
            return render(request, 'tourney/pairing/byebuster_generate.html', {'form': form})
        form = ByebusterGenerateForm(request.POST, tournament=tournament, divisions=divisions)
        if not form.is_valid():
            return render(request, 'tourney/pairing/byebuster_generate.html', {'form': form})
        byebuster_options = form.cleaned_data

    schedules = {}
    errors = []
    pairing_conflicts = []
    byebuster_teams = []
    for division in divisions:
        teams = get_group_teams_for_pairing(tournament, division)
        if len(teams) * tournament.team_prelim_rounds % 2 == 1:
            schedule, schedule_errors, schedule_conflicts, byebuster_team = build_byebuster_prelim_schedule(
                teams,
                tournament.prelim_rounds,
                byebuster_options['counted_rounds'],
                byebuster_options['byebuster_school'],
                byebuster_options['distribution'],
                byebuster_options['lightest_round'],
            )
            if byebuster_team:
                byebuster_teams.append(byebuster_team)
        else:
            schedule, schedule_errors, schedule_conflicts = build_random_prelim_schedule(
                teams,
                tournament.prelim_rounds,
                tournament.team_prelim_rounds,
            )
        if schedule_errors:
            errors.extend(schedule_errors if division is None else [f'{division}: {error}' for error in schedule_errors])
        else:
            schedules[division] = schedule
            for conflict in schedule_conflicts:
                conflict['division'] = division
            pairing_conflicts.extend(schedule_conflicts)
    if errors:
        set_pairing_banner(request, tournament, errors + ['Please create or edit pairings manually.'])
        return redirect('tourney:pairing_index')

    try:
        with transaction.atomic():
            Ballot.objects.filter(round__pairing__tournament=tournament, round__pairing__round_num__lte=tournament.prelim_rounds).delete()
            Pairing.objects.filter(tournament=tournament, round_num__lte=tournament.prelim_rounds).delete()
            Team.objects.filter(user__tournament=tournament).update(byebuster=False)
            for byebuster_team in byebuster_teams:
                Team.objects.filter(pk=byebuster_team.pk).update(byebuster=True)
            assigned_judge_rounds = []
            unassigned_judge_rounds = []
            for round_num in range(1, tournament.prelim_rounds + 1):
                round_objects = []
                pairings_for_round = []
                for division in divisions:
                    pairing = Pairing.objects.create(
                        tournament=tournament,
                        round_num=round_num,
                        division=division,
                        team_submit=True,
                        final_submit=False,
                    )
                    pairings_for_round.append(pairing)
                    letters = get_pairing_letters(division, len(schedules[division][round_num - 1]))
                    for index, (p_team, d_team) in enumerate(schedules[division][round_num - 1]):
                        round_obj = Round.objects.create(
                            pairing=pairing,
                            p_team=p_team,
                            d_team=d_team,
                            courtroom=letters[index],
                        )
                        round_objects.append(round_obj)

                if assign_judges_for_rounds(tournament, round_num, round_objects):
                    assigned_judge_rounds.append(tournament.get_round_label(round_num))
                    for round_obj in round_objects:
                        round_obj.save(update_fields=['presiding_judge', 'scoring_judge', 'extra_judge'])
                    for pairing in pairings_for_round:
                        pairing.final_submit = True
                        pairing.save(update_fields=['final_submit'])
                        sync_ballots_for_pairing(pairing)
                else:
                    unassigned_judge_rounds.append(tournament.get_round_label(round_num))
    except ValidationError as exc:
        set_pairing_banner(request, tournament, [str(exc), 'Please edit pairings manually.'])
        return redirect('tourney:pairing_index')

    messages = ['Preliminary rounds were auto-generated. Review them before publishing.']
    if assigned_judge_rounds and not unassigned_judge_rounds:
        messages.append('Judges were auto-assigned for all generated preliminary rounds.')
    elif assigned_judge_rounds:
        messages.append('Judges were auto-assigned for: ' + ', '.join(assigned_judge_rounds))
        messages.append('Pairings were generated without judges for: ' + ', '.join(unassigned_judge_rounds))
    else:
        messages.append('Pairings were generated without judges. Use Assign Judges after loading judges.')
    if byebuster_teams:
        messages.append(
            'Byebuster team(s): ' + ', '.join(str(team) for team in byebuster_teams)
        )
    set_pairing_banner(
        request,
        tournament,
        messages,
        auto_pairing_conflicts=pairing_conflicts,
    )
    return redirect('tourney:pairing_index')


def sync_ballots_for_pairing(pairing):
    if not pairing.final_submit:
        return
    for round_obj in pairing.rounds.all():
        judges = round_obj.judges
        for judge in judges:
            Ballot.objects.get_or_create(round=round_obj, judge=judge)
        Ballot.objects.filter(round=round_obj).exclude(judge__in=judges).delete()
        round_obj.sync_counted_ballots()


def mark_random_byebuster_exclusion(byebuster_team):
    ballots = list(Ballot.objects.filter(round__pairing__tournament=byebuster_team.user.tournament, submit=True).filter(
        Q(round__p_team=byebuster_team) | Q(round__d_team=byebuster_team)
    ))
    if not ballots:
        return False
    ballot = random.choice(ballots)
    ballot.byebuster_excluded_team = byebuster_team
    ballot.save(update_fields=['byebuster_excluded_team'])
    byebuster_team.save()
    return True


def finalize_pending_byebuster_exclusions(tournament):
    changed = False
    for byebuster_team in Team.objects.filter(user__tournament=tournament, byebuster=True):
        if Ballot.objects.filter(
            round__pairing__tournament=tournament,
            byebuster_excluded_team=byebuster_team,
        ).exists():
            continue
        ballots = list(Ballot.objects.filter(round__pairing__tournament=tournament).filter(
            Q(round__p_team=byebuster_team) | Q(round__d_team=byebuster_team)
        ))
        prelim_ballots = [
            ballot for ballot in ballots
            if not tournament.is_elim_round(ballot.round.pairing.round_num)
        ]
        if prelim_ballots and all(ballot.submit for ballot in prelim_ballots):
            ballot = random.choice(prelim_ballots)
            ballot.byebuster_excluded_team = byebuster_team
            ballot.save(update_fields=['byebuster_excluded_team'])
            byebuster_team.save()
            changed = True
    return changed


def build_round_formset(pairing_capacity, extra_forms):
    return inlineformset_factory(
        Pairing,
        Round,
        form=RoundForm,
        formset=PairingFormSet,
        max_num=pairing_capacity,
        validate_max=True,
        extra=extra_forms,
    )


def has_conflict_errors(formsets):
    keywords = ['conflict', 'judged team', 'played each other', 'supposed to play']
    for formset in formsets:
        for form_errors in formset.errors:
            for value in form_errors.values():
                text = str(value).lower()
                if any(keyword in text for keyword in keywords):
                    return True
        if any(keyword in str(formset.non_form_errors()).lower() for keyword in keywords):
            return True
    return False


def index(request):
    return render(request, 'index.html')


@user_passes_test(lambda u: u.is_staff)
def results(request):
    tournament = request.user.tournament
    if tournament.split_division:
        div1_teams = sorted_break_teams(
            tournament, Team.objects.filter(user__tournament=tournament, division='Disney'))
        div2_teams = sorted_break_teams(
            tournament, Team.objects.filter(user__tournament=tournament, division='Universal'))
        dict = {'teams_ranked': [div1_teams, div2_teams]}
    else:
        teams = sorted_break_teams(tournament)
        dict = {'teams_ranked': teams}
    return render(request, 'tourney/tab/results.html', dict)


@user_passes_test(lambda u: u.is_staff)
def elim_results(request):
    tournament = request.user.tournament
    elim_pairings = Pairing.objects.filter(
        tournament=tournament,
        round_num__gt=tournament.prelim_rounds,
    ).order_by('round_num')
    rounds_by_pairing = []
    for pairing in elim_pairings:
        rounds = []
        for round_obj in pairing.rounds.all().order_by('courtroom'):
            rounds.append({
                'round': round_obj,
                'winner': get_round_winner(round_obj),
            })
        rounds_by_pairing.append((pairing, rounds))
    return render(request, 'tourney/tab/elim_results.html', {'rounds_by_pairing': rounds_by_pairing})


@user_passes_test(lambda u: u.is_staff)
def individual_awards(request):
    tournament = request.user.tournament
    competitors = list(Competitor.objects.filter(team__user__tournament=tournament))
    ranked = []
    for competitor in competitors:
        base_score = competitor.total_score
        if tournament.individual_award_rank_plus_record:
            base_score += competitor.team.total_ballots
        ranked.append((competitor, base_score))
    ranked.sort(key=lambda item: item[1], reverse=True)

    dict = {'ranked': ranked}
    return render(request, 'tourney/tab/individual_awards.html', dict)


@user_passes_test(lambda u: u.is_staff)
def next_pairing(request, round_num):
    tournament = request.user.tournament
    next_round = round_num + 1
    round_title = get_round_title(tournament, next_round)
    if tournament.is_elim_round(next_round):
        elim_pairs = build_elim_pairings(tournament, next_round)
        dict = {
            'next_round': next_round,
            'next_round_title': round_title,
            'divs': ['Break'],
            'teams': [elim_pairs],
            'is_elim': True,
        }
        return render(request, 'tourney/pairing/next_pairing.html', dict)

    if tournament.split_division:
        dict = {'next_round': next_round,
                'next_round_title': round_title,
                'divs': ['Disney', 'Universal'],
                'teams': [build_prelim_pairings(tournament, next_round, 'Disney'),
                          build_prelim_pairings(tournament, next_round, 'Universal')],
                'is_elim': False,
                }
    else:
        dict = {'next_round': next_round,
                'next_round_title': round_title,
                'divs': ['Teams'],
                'teams': [build_prelim_pairings(tournament, next_round)],
                'is_elim': False,
                }
    return render(request, 'tourney/pairing/next_pairing.html', dict)


@user_passes_test(lambda u: u.is_staff)
def pairing_index(request):
    tournament = request.user.tournament
    round_num_lists = sorted(Pairing.objects.filter(
        tournament=tournament).values_list('round_num', flat=True).distinct())
    pairings = []
    for round_num in round_num_lists:
        pairings.append(Pairing.objects.filter(tournament=tournament,
                                               round_num=round_num).order_by('division'))
    if Pairing.objects.filter(tournament=tournament).exists():
        next_round = max([pairing.round_num for pairing in Pairing.objects.filter(
            tournament=tournament)]) + 1
    else:
        next_round = 1
    dict = {
        'pairings': pairings,
        'next_round': next_round,
        'can_add_pairing': next_round <= tournament.total_rounds,
        'next_round_title': get_round_title(tournament, next_round) if next_round <= tournament.total_rounds else None,
        'show_generate_prelims': tournament.randomize_prelims and not Pairing.objects.filter(tournament=tournament).exists(),
    }
    if request.session.get('extra'):
        extra = request.session['extra']
        if extra.get('tournament_id') != tournament.pk:
            request.session.pop('extra', None)
        else:
            dict.update(extra)
    return render(request, 'tourney/pairing/main.html', dict)


@user_passes_test(lambda u: u.is_staff)
def edit_pairing(request, round_num):
    tournament = request.user.tournament
    pairing_capacity = get_pairing_capacity(tournament, round_num)
    waive_conflicts = request.method == "POST" and request.POST.get('waive_conflicts') == '1'
    max_judges = 9 if tournament.is_elim_round(round_num) else tournament.get_max_judges_for_round(round_num)

    if request.user.tournament.split_division and not tournament.is_elim_round(round_num):
        if not Pairing.objects.filter(tournament=tournament, round_num=round_num).exists():
            div1_pairing = Pairing.objects.create(
                tournament=tournament, round_num=round_num, division='Disney')
            div2_pairing = Pairing.objects.create(
                tournament=tournament, round_num=round_num, division='Universal')
        else:
            div1_pairing = Pairing.objects.filter(tournament=tournament,
                round_num=round_num).get(division='Disney')
            div2_pairing = Pairing.objects.filter(tournament=tournament,
                round_num=round_num).get(division='Universal')

        div1_extra_forms = pairing_capacity if not div1_pairing.rounds.exists() else 0
        div2_extra_forms = pairing_capacity if not div2_pairing.rounds.exists() else 0
        Div1RoundFormSet = build_round_formset(pairing_capacity, div1_extra_forms)
        Div2RoundFormSet = build_round_formset(pairing_capacity, div2_extra_forms)

        if not tournament.is_elim_round(round_num):
            if not div1_pairing.rounds.exists():
                for index, (p_team, d_team) in enumerate(build_prelim_pairings(tournament, round_num, 'Disney'), start=1):
                    Round.objects.create(pairing=div1_pairing, p_team=p_team, d_team=d_team,
                                         courtroom=get_pairing_letters('Disney', index)[-1])
            if not div2_pairing.rounds.exists():
                for index, (p_team, d_team) in enumerate(build_prelim_pairings(tournament, round_num, 'Universal'), start=1):
                    Round.objects.create(pairing=div2_pairing, p_team=p_team, d_team=d_team,
                                         courtroom=get_pairing_letters('Universal', index)[-1])

        available_judges_pk = [judge.pk for judge in Judge.objects.all()
                               if judge.get_availability(div1_pairing.round_num)]
        judges = Judge.objects.filter(pk__in=available_judges_pk).order_by(
            '-checkin', '-preside', 'user__username').all()

        if request.method == "POST":
            div1_formset = Div1RoundFormSet(request.POST, request.FILES, prefix='div1', instance=div1_pairing,
                                        form_kwargs={'pairing': div1_pairing, 'other_formset': None, 'request': request,
                                                     'waive_conflicts': waive_conflicts})
            div2_formset = Div2RoundFormSet(request.POST, request.FILES, prefix='div2', instance=div2_pairing,
                                        form_kwargs={'pairing': div2_pairing, 'other_formset': div1_formset, 'request': request,
                                                     'waive_conflicts': waive_conflicts})

            div1_submit_form = PairingSubmitForm(
                request.POST, prefix='div1', instance=div1_pairing)
            div2_submit_form = PairingSubmitForm(
                request.POST, prefix='div2', instance=div2_pairing)

            if div1_submit_form.is_valid():
                div1_submit_form.save()
            if div2_submit_form.is_valid():
                div2_submit_form.save()
            both_true = True
            if div1_formset.is_valid():
                # get courtroom
                actual_round_num = len(div1_formset)
                for form in div1_formset:
                    if form.instance.p_team == None or form.instance.d_team == None:
                        actual_round_num -= 1
                if div1_formset[0].instance.pairing.division == 'Disney':
                    random_choice = get_pairing_letters('Disney', actual_round_num)
                else:
                    random_choice = get_pairing_letters('Universal', actual_round_num)
                for round in Pairing.objects.get(pk=div1_pairing.pk).rounds.all():
                    if round.courtroom != None:
                        random_choice = [
                            label for label in random_choice if label != round.courtroom
                        ]
                random_choice = random.sample(
                    random_choice, len(random_choice))
                for form in div1_formset:
                    if form.instance.p_team != None and form.instance.d_team != None \
                            and form.instance.courtroom == None:
                        form.instance.courtroom = random_choice[0]
                        del (random_choice[0])
                        form.save()
                div1_formset.save()
            else:
                both_true = False

            if div2_formset.is_valid():
                actual_round_num = len(div2_formset)
                for form in div2_formset:
                    if form.instance.p_team == None or form.instance.d_team == None:
                        actual_round_num -= 1
                if div2_formset[0].instance.pairing.division == 'Disney':
                    random_choice = get_pairing_letters('Disney', actual_round_num)
                else:
                    random_choice = get_pairing_letters('Universal', actual_round_num)
                for round in Pairing.objects.get(pk=div2_pairing.pk).rounds.all():
                    if round.courtroom != None:
                        random_choice = [
                            label for label in random_choice if label != round.courtroom
                        ]
                random_choice = random.sample(
                    random_choice, len(random_choice))
                for form in div2_formset:
                    if form.instance.p_team != None and form.instance.d_team != None \
                            and form.instance.courtroom == None:
                        form.instance.courtroom = random_choice[0]
                        del (random_choice[0])
                        form.save()
                div2_formset.save()
            else:
                both_true = False

            pairings = [div1_pairing, div2_pairing]
            for pairing in pairings:
                sync_ballots_for_pairing(pairing)

            if both_true:
                return redirect('tourney:pairing_index')
        else:
            div1_formset = Div1RoundFormSet(instance=div1_pairing, prefix='div1',
                                        form_kwargs={'pairing': div1_pairing,
                                                     'other_formset': None,
                                                     'request': request,
                                                     'waive_conflicts': False})
            div2_formset = Div2RoundFormSet(instance=div2_pairing, prefix='div2',
                                        form_kwargs={'pairing': div2_pairing,
                                                     'other_formset': div1_formset,
                                                     'request': request,
                                                     'waive_conflicts': False})
            div1_submit_form = PairingSubmitForm(
                instance=div1_pairing, prefix='div1')
            div2_submit_form = PairingSubmitForm(
                instance=div2_pairing, prefix='div2')

        return render(request, 'tourney/pairing/edit.html', {'formsets': [div1_formset, div2_formset],
                                                             'submit_forms': [div1_submit_form, div2_submit_form],
                                                             'pairing': div1_pairing,
                                                             'judges': judges,
                                                             'max_judges': max_judges,
                                                             'is_elim': tournament.is_elim_round(round_num),
                                                             'show_waive_conflicts': has_conflict_errors([div1_formset, div2_formset])})
    else:
        if not Pairing.objects.filter(tournament=tournament, round_num=round_num).exists():
            pairing = Pairing.objects.create(
                tournament=tournament, round_num=round_num)
        else:
            pairing = Pairing.objects.get(
                tournament=tournament, round_num=round_num)

        extra_forms = pairing_capacity if not pairing.rounds.exists() else 0
        RoundFormSet = build_round_formset(pairing_capacity, extra_forms)

        if tournament.is_elim_round(round_num) and not pairing.rounds.exists():
            lock_prelim_results(tournament)
            elim_pairs = build_elim_pairings(tournament, round_num)
            letters = random.sample(get_pairing_letters(None, len(elim_pairs)), len(elim_pairs))
            for index, (p_team, d_team) in enumerate(elim_pairs):
                Round.objects.create(
                    pairing=pairing,
                    p_team=p_team,
                    d_team=d_team,
                    courtroom=letters[index],
                )
        elif not tournament.is_elim_round(round_num) and not pairing.rounds.exists():
            for index, (p_team, d_team) in enumerate(build_prelim_pairings(tournament, round_num), start=1):
                Round.objects.create(
                    pairing=pairing,
                    p_team=p_team,
                    d_team=d_team,
                    courtroom=get_pairing_letters(None, index)[-1],
                )

        available_judges_pk = [judge.pk for judge in Judge.objects.filter(user__tournament=tournament)
                               if judge.get_availability(pairing.round_num)]
        judges = Judge.objects.filter(pk__in=available_judges_pk).order_by(
            '-checkin', '-preside', 'user__username').all()

        if request.method == "POST":
            formset = RoundFormSet(request.POST, request.FILES, prefix='div1', instance=pairing,
                                   form_kwargs={'pairing': pairing,
                                                'other_formset': None,
                                                'request': request,
                                                'waive_conflicts': waive_conflicts})
            submit_form = PairingSubmitForm(
                request.POST, prefix='div1', instance=pairing)

            if submit_form.is_valid():
                submit_form.save()

            both_true = True
            if formset.is_valid():
                # get courtroom
                actual_round_num = len(formset)
                for form in formset:
                    if form.instance.p_team == None or form.instance.d_team == None:
                        actual_round_num -= 1

                random_choice = get_pairing_letters(pairing.division, max(pairing_capacity, actual_round_num))[:actual_round_num]

                for round in Pairing.objects.get(pk=pairing.pk).rounds.all():
                    if round.courtroom != None:
                        random_choice = [
                            label for label in random_choice if label != round.courtroom
                        ]
                random_choice = random.sample(
                    random_choice, len(random_choice))
                for form in formset:
                    if form.instance.p_team != None and form.instance.d_team != None \
                            and form.instance.courtroom == None:
                        form.instance.courtroom = random_choice[0]
                        del (random_choice[0])
                        form.save()
                formset.save()
            else:
                both_true = False

            if both_true:
                sync_ballots_for_pairing(pairing)
                return redirect('tourney:pairing_index')
        else:
            formset = RoundFormSet(instance=pairing, prefix='div1',
                                   form_kwargs={'pairing': pairing,
                                                'other_formset': None,
                                                'request': request,
                                                'waive_conflicts': False})
            submit_form = PairingSubmitForm(instance=pairing, prefix='div1')

        return render(request, 'tourney/pairing/edit.html', {'formsets': [formset],
                                                             'submit_forms': [submit_form],
                                                             'pairing': pairing,
                                                             'judges': judges,
                                                             'round_title': get_round_title(tournament, round_num),
                                                             'max_judges': max_judges,
                                                             'is_elim': tournament.is_elim_round(round_num),
                                                             'show_waive_conflicts': has_conflict_errors([formset])})


@user_passes_test(lambda u: u.is_staff)
def delete_pairing(request, round_num):
    errors = []
    tournament = request.user.tournament
    cur_pairing = Pairing.objects.filter(tournament=tournament, round_num=round_num)
    if cur_pairing.exists():
        last_round_num = Pairing.objects.filter(tournament=tournament).order_by('-round_num').values_list('round_num', flat=True).first()
        if last_round_num == round_num:
            Ballot.objects.filter(round__pairing__in=cur_pairing).delete()
            cur_pairing.delete()
            if round_num <= tournament.prelim_rounds:
                Team.objects.filter(user__tournament=tournament).update(
                    byebuster=False,
                    prelim_seed=None,
                    locked_prelim_ballots=None,
                    locked_prelim_cs=None,
                    locked_prelim_pd=None,
                )
        else:
            errors.append('You can only delete the last pairing!')
    set_pairing_banner(request, tournament, errors)
    return redirect('tourney:pairing_index')

    # request, 'tourney/pairing/main.html', {'errors':errors})


@login_required
def view_pairing(request, pk):
    pairing = Pairing.objects.get(pk=pk)
    if not pairing.team_submit:
        context = {}
    else:
        context = {'pairing': [pairing]}
    return render(request, 'tourney/pairing/view.html', context)


@user_passes_test(lambda u: u.is_staff)
def checkin_judges(request, round_num):
    if request.method == "POST":
        form = CheckinJudgeForm(
            request.POST, round_num=round_num, request=request)
        if form.is_valid():
            for judge in form.cleaned_data['checkins']:
                judge.checkin = True
                judge.save()
        return redirect('tourney:pairing_index')
    else:
        form = CheckinJudgeForm(round_num=round_num, request=request)

    return render(request, 'tourney/tab/checkin_judges.html', {'form': form, 'round_num': round_num})


@user_passes_test(lambda u: u.is_staff)
def view_teams(request):
    teams = Team.objects.filter(user__tournament=request.user.tournament)
    return render(request, 'tourney/tab/view_teams.html', {'teams': teams})


@user_passes_test(lambda u: u.is_staff)
def view_judges(request):
    judges = Judge.objects.filter(user__tournament=request.user.tournament)
    return render(request, 'tourney/tab/view_judges.html', {'judges': judges})


@user_passes_test(lambda u: u.is_staff)
def view_individual_judge(request, pk):
    judge = Judge.objects.get(pk=pk)

    if request.method == 'POST':
        user_form = UpdateConflictForm(
            data=request.POST, instance=judge, request=request)
        judge_form = JudgeForm(data=request.POST, instance=judge, request=request)
        if user_form.is_valid():
            user_form.save()
        if judge_form.is_valid():
            judge_form.save()
        return redirect('tourney:view_judges')
    else:
        user_form = UpdateConflictForm(instance=judge, request=request)
        judge_form = JudgeForm(instance=judge, request=request)

    context = {
        'conflict_form': user_form,
        'preference_form': judge_form,
        'selected_conflict_ids': user_form.selected_conflict_ids(),
    }
    return render(request, 'tourney/tab/view_individual_judge.html', context)


@login_required
def view_individual_team(request, pk):
    tournament = request.user.tournament
    # if not Team.objects.filter(user__tournament=tournament, pk=pk).exists():
    #     team = Team.objects.create(user__tournament=tournament)
    # else:
    team = Team.objects.get(user__tournament=tournament, pk=pk)
    if not (request.user.is_team and request.user.team == team) and not request.user.is_staff:
        return HttpResponseNotFound('<h1>Page not found</h1>')
    FormSet = get_team_competitor_formset(tournament)
    roster_locked = tournament.predetermined_speakers and request.user.is_team and not request.user.is_staff

    if request.method == 'POST':
        if roster_locked:
            return HttpResponseForbidden('Team roster is locked because predetermined speakers are enabled.')
        formset = FormSet(request.POST, request.FILES,
                          prefix='competitors', instance=team)
        team_form = TeamForm(data=request.POST, instance=team)
        if formset.is_valid() and team_form.is_valid():
            team_form.save()
            formset.save()
            if request.user.is_staff:
                return redirect('tourney:view_teams')
            else:
                return redirect('index')
    else:
        formset = FormSet(prefix='competitors', instance=team)
        team_form = TeamForm(instance=team)

    if roster_locked:
        lock_form_fields(team_form)
        for form in formset:
            lock_form_fields(form)

    context = {'formset': formset, 'team_form': team_form, 'roster_locked': roster_locked}
    return render(request, 'tourney/tab/view_individual_team.html', context)


@user_passes_test(lambda u: u.is_staff)
def edit_characters(request):
    tournament = request.user.tournament

    FormSet = inlineformset_factory(Tournament, Character, fields=('name', 'side'),
                                    max_num=12, validate_max=True,
                                    extra=6)

    if request.method == 'POST':
        formset = FormSet(request.POST, request.FILES,
                          prefix='characters', instance=tournament)
        if formset.is_valid():
            formset.save()
            return redirect('index')
    else:
        formset = FormSet(prefix='characters', instance=tournament)

    context = {'formset': formset}
    return render(request, 'tourney/tab/edit_characters.html', context)


@user_passes_test(lambda u: u.is_staff)
def delete_individual_judge(request, pk):
    Judge.objects.get(pk=pk).delete()
    return redirect('tourney:view_judges')


@user_passes_test(lambda u: u.is_staff)
def delete_individual_team(request, pk):
    Team.objects.get(pk=pk).delete()
    return redirect('tourney:view_teams')


@user_passes_test(lambda u: u.is_staff)
def clear_checkin(request):
    Judge.objects.update(checkin=False)
    return redirect('tourney:pairing_index')


@user_passes_test(lambda u: u.is_staff)
def checkin_all_judges(request, round_num):
    available_judges_pk = [judge.pk for judge in Judge.objects.filter(user__tournament=request.user.tournament)
                           if judge.get_availability(round_num)]
    Judge.objects.filter(pk__in=available_judges_pk).update(checkin=True)
    return redirect('tourney:pairing_index')


@user_passes_test(lambda u: u.is_staff)
def view_ballot_status(request, pairing_id):
    pairing = Pairing.objects.get(pk=pairing_id)
    finalize_pending_byebuster_exclusions(pairing.tournament)
    ballots = []
    for round in pairing.rounds.all():
        round.sync_counted_ballots()
        round_ballots = list(round.ballots.all())
        show_counted_overlay = bool(round_ballots) and all(ballot.submit for ballot in round_ballots)
        for ballot in round_ballots:
            ballot.show_counted_overlay = show_counted_overlay
            ballots.append(ballot)
    ballots = sorted(ballots, key=lambda x: x.round.courtroom)
    return render(request, 'tourney/tab/view_ballots_status.html', {'ballots': ballots})


@user_passes_test(lambda u: u.is_staff)
def view_spirit_status(request):
    tournament = request.user.tournament
    teams = Team.objects.filter(user__tournament=tournament)
    teams = sorted(teams, key=lambda x: x.spirit_score, reverse=True)
    return render(request, 'tourney/tab/view_spirit_status.html', {'teams': teams})


@user_passes_test(lambda u: u.is_staff)
def add_spirit_forms(request):
    tournament = request.user.tournament
    teams = Team.objects.filter(user__tournament=tournament)
    for team in teams:
        if not Spirit.objects.filter(team=team).exists():
            spirit = Spirit.objects.create(team=team)
    return redirect('tourney:view_spirit_status')


@user_passes_test(lambda u: u.is_staff)
def view_captains_meeting_status(request, pairing_id):
    pairing = Pairing.objects.get(pk=pairing_id)
    captains_meetings = []
    for round in pairing.rounds.all():
        captains_meetings.append(round.captains_meeting)
    captains_meetings = sorted(
        captains_meetings, key=lambda x: x.round.courtroom)
    return render(request, 'tourney/tab/view_captains_meeting_status.html',
                  {'captains_meetings': captains_meetings})

# @login_required
# # def add_conflict(request):
# #     if request.method == 'POST':
# #         form = AddConflictForm(data=request.POST)
# #         if form.is_valid():
# #
# #     return render(request, 'tourney/add_conflict.html', {'form':form})


class TournamentUpdateView(TabOnlyMixin, UpdateView):
    model = Tournament
    form_class = CreateTournamentForm
    template_name = 'tourney/tab/tournament_settings.html'

    def get_object(self, queryset=None):
        return self.request.user.tournament

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.object.predetermined_speakers:
            rounds = Round.objects.filter(pairing__tournament=self.object).select_related(
                'pairing__tournament',
                'p_team',
                'd_team',
            )
            for round_obj in rounds:
                captains_meeting, _ = CaptainsMeeting.objects.get_or_create(round=round_obj)
                round_obj._apply_predetermined_speakers(captains_meeting)
        return response

    def get_success_url(self):
        # if self.request.user.tournament.spirit:
        add_spirit_forms(self.request)
        return reverse_lazy('load_sections')


class ConflictUpdateView(JudgeOnlyMixin, PassRequestToFormViewMixin, UpdateView):
    model = Judge
    template_name = "tourney/add_conflict.html"

    form_class = UpdateConflictForm

    def get_form(self, form_class=None):
        form = super(ConflictUpdateView, self).get_form(form_class)
        form.fields['conflicts'].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['selected_conflict_ids'] = context['form'].selected_conflict_ids()
        return context

    def get_object(self, queryset=None):
        return self.request.user.judge

    success_url = reverse_lazy('index')


class JudgeFriendUpdateView(JudgeOnlyMixin, PassRequestToFormViewMixin, UpdateView):
    model = Judge
    template_name = "utils/generic_form.html"

    form_class = UpdateJudgeFriendForm

    def get_form(self, form_class=None):
        form = super(JudgeFriendUpdateView, self).get_form(form_class)
        form.fields['judge_friends'].required = False
        return form

    def get_object(self, queryset=None):
        return self.request.user.judge

    success_url = reverse_lazy('index')


class JudgePreferenceUpdateView(JudgeOnlyMixin, UpdateView):
    model = Judge
    template_name = "utils/generic_form.html"

    form_class = JudgeForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        kwargs['tournament'] = self.request.user.tournament
        return kwargs

    def get_object(self, queryset=None):
        return self.request.user.judge

    success_url = reverse_lazy('index')


@user_passes_test(lambda u: u.is_team)
def edit_competitor_pronouns(request):
    team = request.user.team
    if request.method == 'POST':
        competitor_pronouns_forms = [CompetitorPronounsForm(request.POST, instance=competitor,
                                                            prefix=competitor.name)
                                     for competitor in team.competitors.all()]
        for form in competitor_pronouns_forms:
            if form.is_valid():
                form.save()

        return redirect('index')
    else:
        competitor_pronouns_forms = [CompetitorPronounsForm(instance=competitor,
                                                            prefix=competitor.name)
                                     for competitor in team.competitors.all()]
    return render(request, 'tourney/competitor_pronouns.html', {
        'team': team,
        'forms': competitor_pronouns_forms
    })


@user_passes_test(lambda u: u.is_staff)
def generate_passwords(request):
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Teams"]
        n = worksheet.max_row
        m = worksheet.max_column
        wb_changed = False
        for i in range(2, n + 1):
            if not worksheet.cell(row=i, column=17).value and worksheet.cell(row=i, column=1).value:
                wb_changed = True
                worksheet.cell(row=i, column=17).value = ''.join(
                    random.choices(string.ascii_letters + string.digits, k=4))
            if not worksheet.cell(row=i, column=16).value and worksheet.cell(row=i, column=1).value:
                wb_changed = True
                team_name = normalize_import_username(worksheet.cell(row=i, column=1).value)
                tournament_prefix = normalize_import_username(request.user.tournament.short_name)
                worksheet.cell(row=i, column=16).value = f'{tournament_prefix}_{team_name}'

        worksheet = wb["Judges"]
        judge_columns = get_judge_sheet_columns(request.user.tournament, worksheet, create_missing=True)
        n = worksheet.max_row
        for i in range(judge_columns['header_row'] + 1, n + 1):
            first_name = worksheet.cell(i, judge_columns['first_name']).value
            last_name = worksheet.cell(i, judge_columns['last_name']).value

            if not worksheet.cell(row=i, column=judge_columns['password']).value:
                wb_changed = True
                worksheet.cell(row=i, column=judge_columns['password']).value = ''.join(
                    random.choices(string.ascii_letters + string.digits, k=4))
            if not worksheet.cell(row=i, column=judge_columns['username']).value and first_name and last_name:
                wb_changed = True
                worksheet.cell(
                    row=i, column=judge_columns['username']).value = get_judge_import_username(
                        request.user.tournament,
                        first_name,
                        last_name,
                    )

        response = HttpResponse(content_type='application/vnd.ms-excel')
        wb.save(response)
        return response


@transaction.non_atomic_requests
@user_passes_test(lambda u: u.is_staff)
def load_teams_and_judges(request):
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        team_list, wb_changed = load_teams_wrapper(request, wb)
        judge_list, wb_changed2 = load_judges_wrapper(request, wb)
        return render(request, 'admin/load_excel.html', {"list": team_list + judge_list})


@transaction.non_atomic_requests
@user_passes_test(lambda u: u.is_staff)
def load_teams(request):
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        response_list, wb_changed = load_teams_wrapper(request, wb)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        wb.save(response)
        return render(request, 'admin/load_excel.html', {"list": response_list})


def load_teams_wrapper(request, wb):
    worksheet = wb["Teams"]
    list = []
    n = worksheet.max_row
    m = worksheet.max_column
    wb_changed = False
    for i in range(2, n + 1):
        team_name = worksheet.cell(i, 1).value
        if not team_name:
            continue

        school = worksheet.cell(i, 2).value
        j = 3
        team_roster = []
        while j <= m and worksheet.cell(i, j).value != None and worksheet.cell(i, j).value != '':
            team_roster.append(worksheet.cell(i, j).value)
            j += 1
        message = ''
        username = worksheet.cell(i, 16).value
        try:
            with transaction.atomic():
                if Team.objects.filter(user__tournament=request.user.tournament, team_name=team_name).exists():
                    team = Team.objects.get(user__tournament=request.user.tournament, team_name=team_name)
                else:
                    team = None

                if team:
                    Team.objects.filter(pk=team.pk).update(
                        team_name=team_name, school=school)
                    team = Team.objects.get(pk=team.pk)
                    message += f' update {team_name} \n'
                else:
                    raw_password = worksheet.cell(i, 17).value
                    if not username:
                        raise ValueError(f'Missing username for team {team_name}')
                    if not raw_password:
                        raise ValueError(f'Missing password for team {team_name}')
                    user = User(username=username, raw_password=raw_password, is_team=True, is_judge=False,
                                tournament=request.user.tournament)
                    user.set_password(raw_password)
                    user.save()
                    team = Team(user=user, team_name=team_name,
                                school=school)
                    team.save()
                    message += f' create {team_name} \n'
                created_roster = []
                updated_roster = []
                for name in team_roster:
                    name = re.sub(r'\([^)]*\)', '', name).strip()
                    if Competitor.objects.filter(team=team, name=name).exists():
                        updated_roster.append(name)
                        Competitor.objects.filter(
                            team=team, name=name).update(team=team, name=name)
                    else:
                        created_roster.append(name)
                        Competitor.objects.create(name=name, team=team)
                if created_roster:
                    str_created_roster = ' , '.join(created_roster)
                    message += f' created roster {str_created_roster} \n'
                if updated_roster:
                    str_updated_roster = ','.join(updated_roster)
                    message += f' updated roster {str_updated_roster} \n'

        except Exception as e:
            message += str(e)
        else:
            message = ' SUCCESS ' + message

        list.append(message)
    return list, wb_changed


def load_judges_wrapper(request, wb):
    worksheet = wb["Judges"]
    list = []
    n = worksheet.max_row
    wb_changed = False
    judge_columns = get_judge_sheet_columns(request.user.tournament, worksheet, create_missing=True)
    seen_usernames = set()
    seen_name_keys = set()
    for i in range(judge_columns['header_row'] + 1, n + 1):
        first_name = normalize_import_value(worksheet.cell(i, judge_columns['first_name']).value)
        last_name = normalize_import_value(worksheet.cell(i, judge_columns['last_name']).value)
        if not first_name and not last_name:
            continue

        if not worksheet.cell(row=i, column=judge_columns['password']).value:
            wb_changed = True
            worksheet.cell(row=i, column=judge_columns['password']).value = ''.join(
                random.choices(string.ascii_letters + string.digits, k=4))
        if not worksheet.cell(row=i, column=judge_columns['username']).value and first_name and last_name:
            wb_changed = True
            worksheet.cell(
                row=i, column=judge_columns['username']).value = get_judge_import_username(
                    request.user.tournament,
                    first_name,
                    last_name,
                )

        username = normalize_import_username(worksheet.cell(i, judge_columns['username']).value)
        if username == None or username == '':
            continue
        if worksheet.cell(i, judge_columns['username']).value != username:
            wb_changed = True
            worksheet.cell(i, judge_columns['username']).value = username

        if last_name == None or last_name == '':
            last_name = ' '
        name_key = get_judge_import_name_key(first_name, last_name)
        if username in seen_usernames or name_key in seen_name_keys:
            list.append(f' SKIPPED duplicate judge row {username} \n')
            continue
        seen_usernames.add(username)
        seen_name_keys.add(name_key)
        raw_password = worksheet.cell(i, judge_columns['password']).value
        preside = worksheet.cell(i, judge_columns['preside']).value
        if preside in ['CIN', 'No preference']:
            preside = 2
        elif preside in ['Y', 'Presiding', 'Yes', 'y', 'YES']:
            preside = 1
        else:
            preside = 0
        availability_by_round = {}
        for round_num, col in judge_columns['rounds'].items():
            availability_by_round[round_num] = bool(col and worksheet.cell(i, col).value in ['y', 'YES', 'Y', 'Yes'])

        message = ''
        try:
            with transaction.atomic():
                user = find_existing_import_user(
                    request.user.tournament,
                    username,
                    first_name,
                    last_name,
                )
                judge = get_user_judge(user) if user else None
                if judge:
                    message += f'update judge {user.username} \n'
                    if user.username != username and not User.objects.filter(username=username).exists():
                        user.username = username
                    user.first_name = first_name
                    user.last_name = last_name
                    user.tournament = request.user.tournament
                    user.save()

                    judge.preside = preside
                    for index, field_name in enumerate(Judge.availability_field_names()):
                        setattr(judge, field_name, availability_by_round.get(index + 1, False))
                    judge.save()
                elif user:
                    message += f'create judge profile for existing account {user.username} \n'
                    if user.username != username and not User.objects.filter(username=username).exists():
                        user.username = username
                    user.first_name = first_name
                    user.last_name = last_name
                    user.is_team = False
                    user.is_judge = True
                    user.tournament = request.user.tournament
                    if raw_password:
                        user.raw_password = raw_password
                        user.set_password(raw_password)
                    user.save()
                    judge = Judge(user=user, preside=preside)
                    for index, field_name in enumerate(Judge.availability_field_names()):
                        setattr(judge, field_name, availability_by_round.get(index + 1, False))
                    judge.save()
                else:
                    message += f'create judge {username} \n'
                    user = User(username=username,
                                first_name=first_name, last_name=last_name,
                                is_team=False, is_judge=True, tournament=request.user.tournament)
                    user.set_password(raw_password)
                    user.save()
                    judge = Judge(user=user, preside=preside)
                    for index, field_name in enumerate(Judge.availability_field_names()):
                        setattr(judge, field_name, availability_by_round.get(index + 1, False))

                    judge.save()

        except Exception as e:
            message += str(e)
        else:
            message = ' SUCCESS ' + message
        list.append(message)
    return list, wb_changed


@transaction.non_atomic_requests
@user_passes_test(lambda u: u.is_staff)
def load_judges(request):
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        response_list, wb_changed = load_judges_wrapper(request, wb)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        wb.save(response)
        if wb_changed:
            return response
        elif DEBUG:
            return render(request, 'admin/load_excel.html', {"list": response_list, })


@user_passes_test(lambda u: u.is_staff)
def load_paradigms(request):
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Paradigms"]
        list = []
        n = worksheet.max_row
        m = worksheet.max_column
        headers = [None]
        for i in range(1, m):
            headers.append(worksheet.cell(1, i).value)

        for i in range(2, n + 1):
            username = worksheet.cell(i, 1).value
            if username == None or username == '':
                continue
            paradigm_items = []
            for j in range(2, worksheet.max_column):
                value = worksheet.cell(i, j).value
                if value:
                    paradigm_items.append((headers[j], value))

            message = ''
            try:
                if not Judge.objects.filter(user__username=username).exists():
                    continue

                if Paradigm.objects.filter(judge__user__username=username).exists():
                    message += f'update judge paradigm {username}'
                    paradigm = Paradigm.objects.get(
                        judge__user__username=username)
                else:
                    paradigm = Paradigm.objects.create(
                        judge=Judge.objects.get(user__username=username))

                for name, value in paradigm_items:
                    if name == 'experience_description':
                        experiences = value.split(',')
                        experiences_actual_vals = []
                        for experience in experiences:
                            experience = experience.strip()
                            # message += str(experience)
                            for (actual_val, display_val) in experience_description_choices:
                                if experience == display_val:
                                    experiences_actual_vals.append(actual_val)
                        message += str(experiences_actual_vals)
                        setattr(paradigm, name, experiences_actual_vals)
                    elif name == 'experience_years':
                        setattr(paradigm, name, int(value))
                    else:
                        try:
                            paradigm_preference_pk = int(name)
                            if ParadigmPreferenceItem.objects.filter(
                                    paradigm=paradigm, paradigm_preference__pk=paradigm_preference_pk).exists():
                                ParadigmPreferenceItem.objects.filter(
                                    paradigm=paradigm, paradigm_preference__pk=paradigm_preference_pk).update(scale=int(value))
                            else:
                                ParadigmPreferenceItem.objects.create(
                                    paradigm=paradigm, paradigm_preference=ParadigmPreference.objects.get(
                                        pk=paradigm_preference_pk),
                                    scale=int(value))
                        except ValueError:
                            setattr(paradigm, name, value)
                paradigm.save()
            except Exception as e:
                message += str(e)
            else:
                message += ' success '
            list.append(message)
        return render(request, 'admin/load_excel.html', {"list": list})


amta_witnesses = {
    'P': ['Ari Felder', 'Aubrey Roy', 'Drew Hubbard', 'Jamie Savchenko'],
    'D': ['Casey Koller', 'Kennedy Heisman', 'R. Moore'],
    'other': ['D.B. Gelfand', 'Mandy Navarra', 'Shannon Shahid'],
}


@user_passes_test(lambda u: u.is_staff)
def load_sections(request):
    tournament = request.user.tournament
    if not Section.objects.filter(tournament=tournament).exists():
        speaker_nums = tournament.wit_nums
        i = 1
        side_choices = {
            'P': tournament.p_choice,
            'D': 'Respondent'
        }
        for side in ['P', 'D']:
            for speaker_num in range(1, speaker_nums + 1):
                section = Section.objects.create(
                    name=f'{side_choices[side]} Speaker {speaker_num}', tournament=tournament)
                SubSection.objects.create(name=f'{side} Speaker {speaker_num} Content',
                                          section=section,
                                          side=side,
                                          role='att',
                                          type='statement',
                                          help_text='Content of Argument',
                                          sequence=i)
                SubSection.objects.create(name=f'{side} Speaker {speaker_num} Extemporaneous',
                                          section=section,
                                          side=side,
                                          role='att',
                                          type='statement',
                                          help_text='Extemporaneous Ability',
                                          sequence=i)
                i += 1
                SubSection.objects.create(name=f'{side} Speaker {speaker_num} Forensics',
                                          section=section,
                                          side=side,
                                          role='att',
                                          type='statement',
                                          help_text='Forensic Skill & Courtroom Demeanor',
                                          sequence=i)
                i += 1
    return redirect('index')


@user_passes_test(lambda u: u.is_staff)
def load_amta_witnesses(request):
    tournament = request.user.tournament
    for side, witnesses in amta_witnesses.items():
        for witness in witnesses:
            Character.objects.create(
                tournament=tournament, name=witness, side=side)
    return redirect('index')


def donate(request):
    return render(request, 'donate.html')


@user_passes_test(lambda u: u.is_staff)
def refresh(request):
    tournament = request.user.tournament
    finalize_pending_byebuster_exclusions(tournament)
    teams = [team for team in Team.objects.filter(user__tournament=tournament)]
    errors = []
    for team in teams:
        team.save()
        # errors.append(f"{team} {team.total_ballots}")
        # for competitor in team.competitors.all():
        #     competitor.save()
    for team in teams:
        team.save()
    if Pairing.objects.filter(tournament=tournament, round_num__gt=tournament.prelim_rounds).exists():
        return redirect('tourney:elim_results')
    return redirect('tourney:results')  # , {'errors': errors}
